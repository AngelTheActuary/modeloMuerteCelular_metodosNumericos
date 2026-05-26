# =============================================================================
#  Modelo de Muerte Celular de Pseudomonas aeruginosa MLS7
# =============================================================================
#
#  La EDO que resolvemos:
#      dX/dt = -kd * X(t)
#  donde X(t) = concentración de bacterias viables (UFC/ml)
#  y kd = constante de muerte (día^-1)
#
#  Solución exacta: X(t) = X0 * exp(-kd * t)
# =============================================================================

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import os

# Para generar las hojas de cálculo con las tablas de resultados
# En Google Colab ya viene instalado; si no, ejecuta: pip install openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# =====================================================================
# PASO 1: DEFINIR LOS DATOS DEL PROBLEMA
# =====================================================================

# Condición inicial: 1.40 x 10^9 UFC/ml en el día 0
X0 = 1.40e9  # UFC/ml

# Datos experimentales de desecación de Pseudomonas aeruginosa MLS7
# Cada par (día, concentración en UFC/ml)
t_exp = np.array([0, 3, 6, 9, 12, 15])          # días
X_exp = np.array([1.40e9, 4.33e8, 1.58e6,        # UFC/ml
                  3.333e5, 5.48e4, 1.87e4])

# =====================================================================
# PASO 2: CALCULAR LA CONSTANTE DE MUERTE kd
# =====================================================================
# Usamos los extremos t=0 y t=15:
#   X(15) = X0 * exp(-kd * 15)
#   => kd = -(1/15) * ln(X(15)/X(0))

X_final = 1.87e4   # X(15) dato experimental
X_inicial = 1.40e9  # X(0)

# Cálculo paso a paso (sin atajos)
razon = X_final / X_inicial          # X(15)/X(0)
logaritmo = np.log(razon)            # ln(X(15)/X(0))  → número negativo
kd = -logaritmo / 15.0               # constante de muerte

print("=" * 65)
print("  CÁLCULO DE LA CONSTANTE DE MUERTE kd")
print("=" * 65)
print(f"  X(0)  = {X_inicial:.4e} UFC/ml")
print(f"  X(15) = {X_final:.4e} UFC/ml")
print(f"  Razón X(15)/X(0) = {razon:.6e}")
print(f"  ln(razón) = {logaritmo:.6f}")
print(f"  kd = {kd:.6f} día⁻¹")
print("=" * 65)

# =====================================================================
# PASO 3: DEFINIR LA EDO (función f)
# =====================================================================
# La EDO es:  dX/dt = f(t, X) = -kd * X
# Es lineal, autónoma (no depende explícitamente de t),
# pero escribimos f(t, X) para mantener la forma general.

def f(t, X):
    """
    Lado derecho de la EDO: dX/dt = -kd * X

    Parámetros:
        t : tiempo (días) — no se usa aquí, pero se incluye por generalidad
        X : concentración de bacterias viables (UFC/ml)

    Retorna:
        dX/dt : tasa de cambio de la población
    """
    return -kd * X

# =====================================================================
# PASO 4: SOLUCIÓN EXACTA (analítica)
# =====================================================================

def solucion_exacta(t):
    """
    Solución analítica de dX/dt = -kd * X
    X(t) = X0 * exp(-kd * t)

    Es la solución de referencia contra la que comparamos los métodos numéricos.
    """
    return X0 * np.exp(-kd * t)

# =====================================================================
# PASO 5: MÉTODO DE EULER EXPLÍCITO (Orden 1)
# =====================================================================
# Fórmula:  X_{n+1} = X_n + h * f(t_n, X_n)
#
# Euler usa UNA SOLA evaluación de la pendiente (al inicio del intervalo).
# Si la función cambia rápido dentro del intervalo, el error es grande.

def metodo_euler(f, t0, X0, h, t_final):
    """
    Método de Euler Explícito para resolver dX/dt = f(t, X).

    Parámetros:
        f       : función f(t, X) del lado derecho de la EDO
        t0      : tiempo inicial
        X0      : condición inicial X(t0)
        h       : tamaño de paso (días)
        t_final : tiempo final de la simulación

    Retorna:
        t_vals : arreglo con los tiempos t0, t0+h, t0+2h, ...
        X_vals : arreglo con las aproximaciones X0, X1, X2, ...
    """
    # Número de pasos necesarios
    N = int((t_final - t0) / h)

    # Crear arreglos vacíos para guardar resultados
    t_vals = np.zeros(N + 1)
    X_vals = np.zeros(N + 1)

    # Condiciones iniciales
    t_vals[0] = t0
    X_vals[0] = X0

    # Iterar paso a paso
    for n in range(N):
        t_n = t_vals[n]       # tiempo actual
        X_n = X_vals[n]       # valor actual de X

        # Euler: avanzar usando la pendiente en (t_n, X_n)
        # X_{n+1} = X_n + h * f(t_n, X_n)
        X_vals[n + 1] = X_n + h * f(t_n, X_n)
        t_vals[n + 1] = t_n + h

    return t_vals, X_vals

# =====================================================================
# PASO 6: MÉTODO DE RUNGE-KUTTA DE ORDEN 2 (Heun)
# =====================================================================
# Fórmula:
#   K1 = f(t_n, X_n)                        → pendiente al INICIO
#   K2 = f(t_n + h, X_n + h * K1)           → pendiente al FINAL (predicción)
#   X_{n+1} = X_n + (h/2) * (K1 + K2)       → promedio de ambas
#
# RK2 evalúa la pendiente en dos puntos: el inicio y el final del intervalo.
# El promedio de ambas pendientes reduce el error a O(h²) por paso.

def metodo_rk2_heun(f, t0, X0, h, t_final):
    """
    Método de Runge-Kutta de Orden 2 (variante de Heun).

    Parámetros:
        f       : función f(t, X) del lado derecho de la EDO
        t0      : tiempo inicial
        X0      : condición inicial
        h       : tamaño de paso
        t_final : tiempo final

    Retorna:
        t_vals, X_vals : arreglos con tiempos y aproximaciones
    """
    N = int((t_final - t0) / h)

    t_vals = np.zeros(N + 1)
    X_vals = np.zeros(N + 1)

    t_vals[0] = t0
    X_vals[0] = X0

    for n in range(N):
        t_n = t_vals[n]
        X_n = X_vals[n]

        # Pendiente al inicio del intervalo
        K1 = f(t_n, X_n)

        # Pendiente al final del intervalo (usando predicción de Euler)
        K2 = f(t_n + h, X_n + h * K1)

        # Promedio de ambas pendientes
        # X_{n+1} = X_n + (h/2) * (K1 + K2)
        X_vals[n + 1] = X_n + (h / 2.0) * (K1 + K2)
        t_vals[n + 1] = t_n + h

    return t_vals, X_vals

# =====================================================================
# PASO 7: MÉTODO DE RUNGE-KUTTA DE ORDEN 4 (RK4)
# =====================================================================
# Fórmula:
#   k1 = h * f(t_n,         X_n)              → pendiente al INICIO
#   k2 = h * f(t_n + h/2,   X_n + k1/2)       → pendiente a la MITAD (1er intento)
#   k3 = h * f(t_n + h/2,   X_n + k2/2)       → pendiente a la MITAD (2do intento, corregida)
#   k4 = h * f(t_n + h,     X_n + k3)          → pendiente al FINAL
#
#   X_{n+1} = X_n + (1/6) * (k1 + 2*k2 + 2*k3 + k4)
#
# RK4 evalúa la pendiente en 4 puntos del intervalo y hace un
# promedio ponderado. Error local de truncamiento: O(h^5).

def metodo_rk4(f, t0, X0, h, t_final):
    """
    Método de Runge-Kutta de Orden 4 (clásico).

    Este es el método más preciso de los tres implementados.
    Usa 4 evaluaciones de f por paso para lograr error O(h^4) global.

    Parámetros:
        f       : función f(t, X) del lado derecho de la EDO
        t0      : tiempo inicial
        X0      : condición inicial
        h       : tamaño de paso
        t_final : tiempo final

    Retorna:
        t_vals, X_vals : arreglos con tiempos y aproximaciones
    """
    N = int((t_final - t0) / h)

    t_vals = np.zeros(N + 1)
    X_vals = np.zeros(N + 1)

    t_vals[0] = t0
    X_vals[0] = X0

    for n in range(N):
        t_n = t_vals[n]
        X_n = X_vals[n]

        # --- Pendiente 1: al inicio del intervalo ---
        k1 = h * f(t_n, X_n)

        # --- Pendiente 2: a la mitad, usando k1 para predecir ---
        k2 = h * f(t_n + h / 2.0, X_n + k1 / 2.0)

        # --- Pendiente 3: a la mitad, usando k2 para corregir ---
        k3 = h * f(t_n + h / 2.0, X_n + k2 / 2.0)

        # --- Pendiente 4: al final del intervalo, usando k3 ---
        k4 = h * f(t_n + h, X_n + k3)

        # --- Promedio ponderado: 1/6 * (k1 + 2*k2 + 2*k3 + k4) ---
        # Las pendientes del medio (k2, k3) tienen peso doble
        X_vals[n + 1] = X_n + (1.0 / 6.0) * (k1 + 2.0 * k2 + 2.0 * k3 + k4)
        t_vals[n + 1] = t_n + h

    return t_vals, X_vals

# =====================================================================
# PASO 8: EJECUTAR TODOS LOS MÉTODOS
# =====================================================================

h = 1.0          # Paso de tiempo: 1 día
t_final = 15.0   # Simular hasta el día 15

print("\n" + "=" * 65)
print("  ANÁLISIS DE ESTABILIDAD")
print("=" * 65)
print(f"  kd * h = {kd:.6f} × {h} = {kd * h:.4f}")
print(f"  Límite de estabilidad Euler/RK2: kd*h < 2.0  →  {'ESTABLE ✓' if kd*h < 2.0 else 'INESTABLE ✗'}")
print(f"  Límite de estabilidad RK4:       kd*h < 2.785 →  {'ESTABLE ✓' if kd*h < 2.785 else 'INESTABLE ✗'}")
print("=" * 65)

# Ejecutar cada método
t_euler, X_euler = metodo_euler(f, 0, X0, h, t_final)
t_rk2,   X_rk2   = metodo_rk2_heun(f, 0, X0, h, t_final)
t_rk4,   X_rk4   = metodo_rk4(f, 0, X0, h, t_final)

# Solución exacta en una malla fina (para graficar curva suave)
t_fino = np.linspace(0, t_final, 500)
X_exacta_fino = solucion_exacta(t_fino)

# Solución exacta en los puntos de los métodos numéricos (para comparar)
X_exacta_euler = solucion_exacta(t_euler)
X_exacta_rk2   = solucion_exacta(t_rk2)
X_exacta_rk4   = solucion_exacta(t_rk4)

# =====================================================================
# PASO 9: IMPRIMIR TABLAS DE RESULTADOS
# =====================================================================

# --- Tabla RK4 detallada (como en el documento) ---
print("\n" + "=" * 80)
print("  TABLA DE RESULTADOS: RUNGE-KUTTA 4 vs SOLUCIÓN EXACTA")
print("=" * 80)
print(f"  {'n':>3}  {'t (día)':>8}  {'X_RK4 (UFC/ml)':>18}  {'X_exacta':>18}  {'Error (%)':>10}")
print("-" * 80)

for n in range(len(t_rk4)):
    t = t_rk4[n]
    xrk4 = X_rk4[n]
    xex = X_exacta_rk4[n]
    # Error relativo porcentual
    if xex != 0:
        error = abs(xrk4 - xex) / abs(xex) * 100.0
    else:
        error = 0.0
    print(f"  {n:3d}  {t:8.1f}  {xrk4:18.6e}  {xex:18.6e}  {error:10.3f}")

print("=" * 80)

# --- Tabla comparativa en los puntos experimentales ---
print("\n" + "=" * 95)
print("  COMPARACIÓN EN PUNTOS EXPERIMENTALES")
print("=" * 95)
print(f"  {'t':>4}  {'X_exp':>14}  {'X_exacta':>14}  {'X_Euler':>14}  {'X_RK2':>14}  {'X_RK4':>14}")
print("-" * 95)

for i, t in enumerate(t_exp):
    idx = int(t)  # índice en los arreglos (porque h=1 y t empieza en 0)
    print(f"  {t:4.0f}  {X_exp[i]:14.4e}  {solucion_exacta(t):14.4e}  "
          f"{X_euler[idx]:14.4e}  {X_rk2[idx]:14.4e}  {X_rk4[idx]:14.4e}")

print("=" * 95)

# =====================================================================
# PASO 9.5: GENERAR HOJAS DE CÁLCULO (.xlsx) CON LAS TABLAS
# =====================================================================
# Creamos un archivo Excel con 3 hojas:
#   Hoja 1: Tabla RK4 vs Solución Exacta (16 filas, paso a paso)
#   Hoja 2: Comparación de los 3 métodos en puntos experimentales
#   Hoja 3: Resumen de parámetros y análisis de estabilidad

# Crear directorio para guardar gráficas y hojas de cálculo
output_dir = "PROYECTO FINAL"
os.makedirs(output_dir, exist_ok=True)

wb = Workbook()

# ---- Estilos reutilizables ----
# Encabezado: fondo azul oscuro, texto blanco, negrita
estilo_encabezado = Font(bold=True, color="FFFFFF", size=11, name="Arial")
relleno_encabezado = PatternFill("solid", fgColor="1F3864")
alineacion_centro = Alignment(horizontal="center", vertical="center")
alineacion_derecha = Alignment(horizontal="right", vertical="center")
borde_fino = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
# Fuente para datos numéricos
fuente_datos = Font(name="Arial", size=10)
# Fuente para títulos de hoja
fuente_titulo = Font(bold=True, size=14, name="Arial", color="1F3864")
# Relleno alternado para filas (gris claro)
relleno_alterno = PatternFill("solid", fgColor="D6E4F0")

# ==================================================================
# HOJA 1: RK4 vs Solución Exacta
# ==================================================================
hoja1 = wb.active
hoja1.title = "RK4 vs Exacta"

# Título de la hoja
hoja1.merge_cells("A1:E1")
hoja1["A1"] = "RUNGE-KUTTA 4 vs SOLUCIÓN EXACTA — Muerte Celular de P. aeruginosa"
hoja1["A1"].font = fuente_titulo
hoja1["A1"].alignment = Alignment(horizontal="center")

# Parámetros del modelo en la fila 2
hoja1.merge_cells("A2:E2")
hoja1["A2"] = f"kd = {kd:.6f} día⁻¹  |  X₀ = {X0:.2e} UFC/ml  |  h = {h} día  |  kd·h = {kd*h:.4f}"
hoja1["A2"].font = Font(name="Arial", size=9, italic=True, color="555555")
hoja1["A2"].alignment = Alignment(horizontal="center")

# Encabezados (fila 4)
encabezados_h1 = ["n", "t (día)", "X_RK4 (UFC/ml)", "X_exacta (UFC/ml)", "Error Relativo (%)"]
for col_idx, titulo in enumerate(encabezados_h1, start=1):
    celda = hoja1.cell(row=4, column=col_idx, value=titulo)
    celda.font = estilo_encabezado
    celda.fill = relleno_encabezado
    celda.alignment = alineacion_centro
    celda.border = borde_fino

# Datos (fila 5 en adelante)
for n in range(len(t_rk4)):
    fila = 5 + n
    t_val = t_rk4[n]
    xrk4 = X_rk4[n]
    xex = X_exacta_rk4[n]
    error_val = abs(xrk4 - xex) / abs(xex) * 100.0 if xex != 0 else 0.0

    hoja1.cell(row=fila, column=1, value=n).font = fuente_datos
    hoja1.cell(row=fila, column=2, value=t_val).font = fuente_datos
    hoja1.cell(row=fila, column=3, value=xrk4).font = fuente_datos
    hoja1.cell(row=fila, column=4, value=xex).font = fuente_datos
    hoja1.cell(row=fila, column=5, value=round(error_val, 3)).font = fuente_datos

    # Aplicar bordes y alineación a cada celda de la fila
    for col_idx in range(1, 6):
        celda = hoja1.cell(row=fila, column=col_idx)
        celda.border = borde_fino
        celda.alignment = alineacion_centro
        # Filas alternas con color de fondo
        if n % 2 == 1:
            celda.fill = relleno_alterno

    # Formato científico para columnas 3 y 4 (las concentraciones)
    hoja1.cell(row=fila, column=3).number_format = '0.000000E+00'
    hoja1.cell(row=fila, column=4).number_format = '0.000000E+00'
    hoja1.cell(row=fila, column=5).number_format = '0.000'

# Ajustar anchos de columna
hoja1.column_dimensions['A'].width = 6
hoja1.column_dimensions['B'].width = 10
hoja1.column_dimensions['C'].width = 22
hoja1.column_dimensions['D'].width = 22
hoja1.column_dimensions['E'].width = 20

# ==================================================================
# HOJA 2: Comparación en puntos experimentales
# ==================================================================
hoja2 = wb.create_sheet("Comparación Métodos")

# Título
hoja2.merge_cells("A1:F1")
hoja2["A1"] = "COMPARACIÓN: Datos Experimentales vs Métodos Numéricos"
hoja2["A1"].font = fuente_titulo
hoja2["A1"].alignment = Alignment(horizontal="center")

hoja2.merge_cells("A2:F2")
hoja2["A2"] = "Pseudomonas aeruginosa MLS7 — Desecación 20 min"
hoja2["A2"].font = Font(name="Arial", size=9, italic=True, color="555555")
hoja2["A2"].alignment = Alignment(horizontal="center")

# Encabezados
encabezados_h2 = ["t (día)", "X_exp (UFC/ml)", "X_exacta (UFC/ml)",
                  "X_Euler (UFC/ml)", "X_RK2 (UFC/ml)", "X_RK4 (UFC/ml)"]
for col_idx, titulo in enumerate(encabezados_h2, start=1):
    celda = hoja2.cell(row=4, column=col_idx, value=titulo)
    celda.font = estilo_encabezado
    celda.fill = relleno_encabezado
    celda.alignment = alineacion_centro
    celda.border = borde_fino

# Datos
for i, t_val in enumerate(t_exp):
    fila = 5 + i
    idx = int(t_val)

    hoja2.cell(row=fila, column=1, value=t_val).font = fuente_datos
    hoja2.cell(row=fila, column=2, value=X_exp[i]).font = fuente_datos
    hoja2.cell(row=fila, column=3, value=solucion_exacta(t_val)).font = fuente_datos
    hoja2.cell(row=fila, column=4, value=X_euler[idx]).font = fuente_datos
    hoja2.cell(row=fila, column=5, value=X_rk2[idx]).font = fuente_datos
    hoja2.cell(row=fila, column=6, value=X_rk4[idx]).font = fuente_datos

    for col_idx in range(1, 7):
        celda = hoja2.cell(row=fila, column=col_idx)
        celda.border = borde_fino
        celda.alignment = alineacion_centro
        if i % 2 == 1:
            celda.fill = relleno_alterno

    # Formato científico para columnas 2 a 6
    for col_idx in range(2, 7):
        hoja2.cell(row=fila, column=col_idx).number_format = '0.00E+00'

# Ajustar anchos
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
    hoja2.column_dimensions[col_letter].width = 20
hoja2.column_dimensions['A'].width = 10

# ==================================================================
# HOJA 3: Errores relativos de cada método en t = 0, 3, 6, ..., 15
# ==================================================================
hoja3 = wb.create_sheet("Errores Relativos")

hoja3.merge_cells("A1:D1")
hoja3["A1"] = "ERROR RELATIVO (%) DE CADA MÉTODO vs SOLUCIÓN EXACTA"
hoja3["A1"].font = fuente_titulo
hoja3["A1"].alignment = Alignment(horizontal="center")

encabezados_h3 = ["t (día)", "Error Euler (%)", "Error RK2 (%)", "Error RK4 (%)"]
for col_idx, titulo in enumerate(encabezados_h3, start=1):
    celda = hoja3.cell(row=3, column=col_idx, value=titulo)
    celda.font = estilo_encabezado
    celda.fill = relleno_encabezado
    celda.alignment = alineacion_centro
    celda.border = borde_fino

for i, t_val in enumerate(t_exp):
    fila = 4 + i
    idx = int(t_val)
    exacto = solucion_exacta(t_val)

    err_euler = abs(X_euler[idx] - exacto) / abs(exacto) * 100 if exacto != 0 else 0
    err_rk2   = abs(X_rk2[idx] - exacto) / abs(exacto) * 100 if exacto != 0 else 0
    err_rk4   = abs(X_rk4[idx] - exacto) / abs(exacto) * 100 if exacto != 0 else 0

    hoja3.cell(row=fila, column=1, value=t_val).font = fuente_datos
    hoja3.cell(row=fila, column=2, value=round(err_euler, 2)).font = fuente_datos
    hoja3.cell(row=fila, column=3, value=round(err_rk2, 2)).font = fuente_datos
    hoja3.cell(row=fila, column=4, value=round(err_rk4, 2)).font = fuente_datos

    for col_idx in range(1, 5):
        celda = hoja3.cell(row=fila, column=col_idx)
        celda.border = borde_fino
        celda.alignment = alineacion_centro
        if i % 2 == 1:
            celda.fill = relleno_alterno

    # Formato de porcentaje con 2 decimales
    for col_idx in range(2, 5):
        hoja3.cell(row=fila, column=col_idx).number_format = '0.00'

# Resumen al final
fila_resumen = 4 + len(t_exp) + 1
hoja3.cell(row=fila_resumen, column=1, value="Orden:").font = Font(bold=True, name="Arial")
hoja3.cell(row=fila_resumen, column=2, value="O(h)").font = Font(bold=True, name="Arial", color="DA3633")
hoja3.cell(row=fila_resumen, column=3, value="O(h²)").font = Font(bold=True, name="Arial", color="F78166")
hoja3.cell(row=fila_resumen, column=4, value="O(h⁴)").font = Font(bold=True, name="Arial", color="1F6FEB")

for col_letter in ['A', 'B', 'C', 'D']:
    hoja3.column_dimensions[col_letter].width = 18
hoja3.column_dimensions['A'].width = 10

# ==================================================================
# GUARDAR EL ARCHIVO EXCEL
# ==================================================================
ruta_excel = os.path.join(output_dir, "Tablas_Metodos_Numericos.xlsx")
wb.save(ruta_excel)
print(f"\n  [✓] Hoja de cálculo guardada en: {ruta_excel}")
print(f"      → Hoja 1: RK4 vs Exacta (16 filas paso a paso)")
print(f"      → Hoja 2: Comparación de los 3 métodos en puntos experimentales")
print(f"      → Hoja 3: Errores relativos (%)")

# =====================================================================
# PASO 10: CONFIGURAR ESTILO DE GRÁFICAS
# =====================================================================
# Usamos un estilo oscuro para mayor contraste visual:
# fondo negro/gris oscuro, líneas brillantes, tipografía clara.

plt.rcParams.update({
    'figure.facecolor': '#0d1117',      # fondo de la figura (negro azulado)
    'axes.facecolor': '#161b22',        # fondo del área de gráfica
    'axes.edgecolor': '#30363d',        # borde de los ejes
    'axes.labelcolor': '#e6edf3',       # color de etiquetas
    'text.color': '#e6edf3',           # color de texto general
    'xtick.color': '#8b949e',          # color de marcas del eje X
    'ytick.color': '#8b949e',          # color de marcas del eje Y
    'grid.color': '#21262d',           # color de la cuadrícula
    'grid.alpha': 0.8,
    'legend.facecolor': '#161b22',     # fondo de la leyenda
    'legend.edgecolor': '#30363d',
    'font.size': 11,
    'axes.titlesize': 14,
    'axes.labelsize': 12,
})

# Colores para cada método
COLOR_EXACTA = '#ffd700'       # Dorado (solución de referencia)
COLOR_RK4    = '#58a6ff'       # Azul brillante
COLOR_RK2    = '#f78166'       # Naranja
COLOR_EULER  = '#da3633'       # Rojo
COLOR_EXP    = '#3fb950'       # Verde (datos experimentales)

# Crear directorio para guardar gráficas y hojas de cálculo
output_dir = "PROYECTO FINAL"
os.makedirs(output_dir, exist_ok=True)

# =====================================================================
# GRÁFICA 1: Comparación de todos los métodos (escala logarítmica)
# =====================================================================
# Usamos escala logarítmica en Y porque X varía de 10^9 a 10^4,
# y en escala lineal los valores pequeños se aplastan contra cero.

fig1, ax1 = plt.subplots(figsize=(12, 7))

# Solución exacta (curva suave)
ax1.plot(t_fino, X_exacta_fino, color=COLOR_EXACTA, linewidth=2.5,
         label='Solución Exacta: $X_0 e^{-k_d t}$', zorder=3)

# Método RK4 (el mejor)
ax1.plot(t_rk4, X_rk4, 's-', color=COLOR_RK4, linewidth=1.5,
         markersize=4, label=f'RK4 (h={h})', alpha=0.9, zorder=4)

# Método RK2 (Heun)
ax1.plot(t_rk2, X_rk2, '^--', color=COLOR_RK2, linewidth=1.5,
         markersize=5, label=f'RK2 Heun (h={h})', alpha=0.85, zorder=2)

# Método Euler
ax1.plot(t_euler, X_euler, 'v:', color=COLOR_EULER, linewidth=1.5,
         markersize=5, label=f'Euler (h={h})', alpha=0.85, zorder=2)

# Datos experimentales (puntos grandes y visibles)
ax1.scatter(t_exp, X_exp, color=COLOR_EXP, s=120, marker='D',
            edgecolors='white', linewidths=1.5, zorder=5,
            label='Datos Experimentales')

# Escala logarítmica en el eje Y
ax1.set_yscale('log')

# Etiquetas y título
ax1.set_xlabel('Tiempo (días)', fontsize=13, fontweight='bold')
ax1.set_ylabel('Concentración X (UFC/ml)', fontsize=13, fontweight='bold')
ax1.set_title('Muerte Celular de Pseudomonas aeruginosa MLS7\n'
              'Comparación de Métodos Numéricos vs Solución Exacta',
              fontsize=15, fontweight='bold', pad=15)

# Leyenda, cuadrícula, límites
ax1.legend(loc='upper right', fontsize=10, framealpha=0.9)
ax1.grid(True, which='both', alpha=0.3)
ax1.set_xlim(-0.5, 15.5)
ax1.set_ylim(1e-1, 5e9)

# Anotación: kd
ax1.text(0.02, 0.02,
         f'$k_d$ = {kd:.6f} día$^{{-1}}$\n'
         f'$k_d \\cdot h$ = {kd*h:.4f}\n'
         f'Paso h = {h} día',
         transform=ax1.transAxes, fontsize=9,
         verticalalignment='bottom',
         bbox=dict(boxstyle='round,pad=0.5', facecolor='#21262d',
                   edgecolor='#58a6ff', alpha=0.9))

plt.tight_layout()
ruta_fig1 = os.path.join(output_dir, "01_comparacion_metodos_log.png")
fig1.savefig(ruta_fig1, dpi=200, bbox_inches='tight')
plt.show()
print(f"  [✓] Gráfica 1 guardada en: {ruta_fig1}")

# =====================================================================
# GRÁFICA 2: Error relativo de cada método respecto a la solución exacta
# =====================================================================
# El error relativo mide qué tan lejos está cada método de la verdad.
# Error(%) = |X_numérico - X_exacto| / |X_exacto| * 100

# Calcular errores (evitando dividir por cero en t=0)
error_euler = np.zeros(len(t_euler))
error_rk2   = np.zeros(len(t_rk2))
error_rk4   = np.zeros(len(t_rk4))

for i in range(len(t_euler)):
    exacto = solucion_exacta(t_euler[i])
    if exacto != 0:
        error_euler[i] = abs(X_euler[i] - exacto) / abs(exacto) * 100.0
        error_rk2[i]   = abs(X_rk2[i] - exacto) / abs(exacto) * 100.0
        error_rk4[i]   = abs(X_rk4[i] - exacto) / abs(exacto) * 100.0

fig2, ax2 = plt.subplots(figsize=(12, 6))

ax2.plot(t_euler, error_euler, 'v-', color=COLOR_EULER, linewidth=2,
         markersize=5, label='Euler (orden 1)', alpha=0.9)
ax2.plot(t_rk2, error_rk2, '^-', color=COLOR_RK2, linewidth=2,
         markersize=5, label='RK2 Heun (orden 2)', alpha=0.9)
ax2.plot(t_rk4, error_rk4, 's-', color=COLOR_RK4, linewidth=2,
         markersize=5, label='RK4 (orden 4)', alpha=0.9)

ax2.set_xlabel('Tiempo (días)', fontsize=13, fontweight='bold')
ax2.set_ylabel('Error Relativo (%)', fontsize=13, fontweight='bold')
ax2.set_title('Error Relativo de cada Método vs Solución Exacta',
              fontsize=15, fontweight='bold', pad=15)

ax2.legend(loc='upper left', fontsize=11, framealpha=0.9)
ax2.grid(True, alpha=0.3)
ax2.set_xlim(-0.5, 15.5)

# Zona de confort: error < 10%
ax2.axhspan(0, 10, color=COLOR_RK4, alpha=0.07, label='_')
ax2.text(14.5, 5, 'Zona\naceptable\n(<10%)', fontsize=8,
         ha='right', color=COLOR_RK4, alpha=0.7)

plt.tight_layout()
ruta_fig2 = os.path.join(output_dir, "02_error_relativo.png")
fig2.savefig(ruta_fig2, dpi=200, bbox_inches='tight')
plt.show()
print(f"  [✓] Gráfica 2 guardada en: {ruta_fig2}")

# =====================================================================
# GRÁFICA 3: Factor de amplificación R(z) — Dominio de estabilidad
# =====================================================================
# Para la EDO dX/dt = -lambda*X, definimos z = -lambda*h.
# Cada método tiene un factor de amplificación R(z) tal que X_{n+1} = R(z)*X_n.
# El método es ESTABLE si |R(z)| < 1.
#
# Euler:  R(z) = 1 + z
# RK2:    R(z) = 1 + z + z²/2
# RK4:    R(z) = 1 + z + z²/2 + z³/6 + z⁴/24
# Exacto: R(z) = e^z

z_vals = np.linspace(-4, 1, 1000)

# Factor de amplificación de cada método
R_euler = 1 + z_vals
R_rk2   = 1 + z_vals + z_vals**2 / 2.0
R_rk4   = 1 + z_vals + z_vals**2 / 2.0 + z_vals**3 / 6.0 + z_vals**4 / 24.0
R_exact = np.exp(z_vals)

fig3, ax3 = plt.subplots(figsize=(12, 6))

ax3.plot(z_vals, np.abs(R_exact), color=COLOR_EXACTA, linewidth=2.5,
         label='$|e^z|$ (exacto)', zorder=3)
ax3.plot(z_vals, np.abs(R_rk4), '--', color=COLOR_RK4, linewidth=2,
         label='$|R_{RK4}(z)|$', zorder=4)
ax3.plot(z_vals, np.abs(R_rk2), '--', color=COLOR_RK2, linewidth=2,
         label='$|R_{RK2}(z)|$')
ax3.plot(z_vals, np.abs(R_euler), '--', color=COLOR_EULER, linewidth=2,
         label='$|R_{Euler}(z)|$')

# Línea de estabilidad |R| = 1
ax3.axhline(y=1.0, color='white', linewidth=1, linestyle=':', alpha=0.5)
ax3.text(-3.8, 1.05, '|R| = 1 (frontera de estabilidad)',
         fontsize=9, color='white', alpha=0.7)

# Marcar z = -kd*h para nuestro problema
z_nuestro = -kd * h
ax3.axvline(x=z_nuestro, color=COLOR_EXP, linewidth=1.5, linestyle='--', alpha=0.7)
ax3.text(z_nuestro + 0.05, 1.8,
         f'z = $-k_d h$ = {z_nuestro:.3f}',
         fontsize=10, color=COLOR_EXP, fontweight='bold')

ax3.set_xlabel('z = $-k_d \\cdot h$', fontsize=13, fontweight='bold')
ax3.set_ylabel('|R(z)| (factor de amplificación)', fontsize=13, fontweight='bold')
ax3.set_title('Dominio de Estabilidad Absoluta\n'
              'El método es estable donde |R(z)| < 1',
              fontsize=15, fontweight='bold', pad=15)

ax3.legend(loc='upper right', fontsize=11, framealpha=0.9)
ax3.grid(True, alpha=0.3)
ax3.set_xlim(-4, 1)
ax3.set_ylim(0, 3)

# Sombrear zona estable
ax3.axhspan(0, 1, color=COLOR_RK4, alpha=0.05)

plt.tight_layout()
ruta_fig3 = os.path.join(output_dir, "03_estabilidad.png")
fig3.savefig(ruta_fig3, dpi=200, bbox_inches='tight')
plt.show()
print(f"  [✓] Gráfica 3 guardada en: {ruta_fig3}")

# =====================================================================
# GRÁFICA 4: RK4 paso a paso — Visualización de las 4 pendientes (k1..k4)
# =====================================================================
# Mostramos gráficamente cómo RK4 calcula las 4 pendientes
# en la primera iteración (n=0, de t=0 a t=1).

fig4, ax4 = plt.subplots(figsize=(12, 7))

# Curva exacta de fondo
t_zoom = np.linspace(0, 2, 300)
ax4.plot(t_zoom, solucion_exacta(t_zoom), color=COLOR_EXACTA,
         linewidth=2, label='Solución exacta', alpha=0.7)

# Valores para la primera iteración
t_n = 0.0
X_n = X0

# k1: pendiente al inicio
k1_val = h * f(t_n, X_n)
pendiente_k1 = f(t_n, X_n)

# k2: pendiente a la mitad (usando k1)
X_k2 = X_n + k1_val / 2.0
k2_val = h * f(t_n + h/2, X_k2)
pendiente_k2 = f(t_n + h/2, X_k2)

# k3: pendiente a la mitad (usando k2)
X_k3 = X_n + k2_val / 2.0
k3_val = h * f(t_n + h/2, X_k3)
pendiente_k3 = f(t_n + h/2, X_k3)

# k4: pendiente al final (usando k3)
X_k4 = X_n + k3_val
k4_val = h * f(t_n + h, X_k4)
pendiente_k4 = f(t_n + h, X_k4)

# Resultado RK4
X_rk4_paso = X_n + (1.0/6.0) * (k1_val + 2*k2_val + 2*k3_val + k4_val)

# Dibujar las rectas tangentes de cada pendiente
dt_linea = 0.4  # extensión de cada línea tangente

# k1 (desde t=0, X=X_n)
t_k1 = np.array([t_n - dt_linea/2, t_n + dt_linea/2])
ax4.plot(t_k1, X_n + pendiente_k1 * (t_k1 - t_n),
         color=COLOR_EULER, linewidth=2.5, label=f'$k_1$ = {k1_val:.3e}')

# k2 (desde t=0.5, X=X_k2)
t_k2 = np.array([0.5 - dt_linea/2, 0.5 + dt_linea/2])
ax4.plot(t_k2, X_k2 + pendiente_k2 * (t_k2 - 0.5),
         color=COLOR_RK2, linewidth=2.5, label=f'$k_2$ = {k2_val:.3e}')

# k3 (desde t=0.5, X=X_k3)
ax4.plot(t_k2, X_k3 + pendiente_k3 * (t_k2 - 0.5),
         color='#bc8cff', linewidth=2.5, label=f'$k_3$ = {k3_val:.3e}')

# k4 (desde t=1, X=X_k4)
t_k4 = np.array([1.0 - dt_linea/2, 1.0 + dt_linea/2])
ax4.plot(t_k4, X_k4 + pendiente_k4 * (t_k4 - 1.0),
         color=COLOR_EXP, linewidth=2.5, label=f'$k_4$ = {k4_val:.3e}')

# Puntos clave
ax4.scatter([0], [X_n], color='white', s=150, zorder=5, edgecolors=COLOR_EXACTA, linewidths=2)
ax4.scatter([1], [X_rk4_paso], color=COLOR_RK4, s=150, zorder=5,
            edgecolors='white', linewidths=2, label=f'$X_1^{{RK4}}$ = {X_rk4_paso:.3e}')
ax4.scatter([1], [solucion_exacta(1)], color=COLOR_EXACTA, s=100, zorder=5,
            marker='*', edgecolors='white', linewidths=1)

ax4.set_xlabel('Tiempo (días)', fontsize=13, fontweight='bold')
ax4.set_ylabel('X (UFC/ml)', fontsize=13, fontweight='bold')
ax4.set_title('RK4: Las 4 Pendientes en la Iteración n=0\n'
              'Evaluación de f en 4 puntos del intervalo [0, 1]',
              fontsize=15, fontweight='bold', pad=15)

ax4.legend(loc='upper right', fontsize=9, framealpha=0.9)
ax4.grid(True, alpha=0.3)

# Fórmula
ax4.text(0.02, 0.02,
         '$X_{n+1} = X_n + \\frac{1}{6}(k_1 + 2k_2 + 2k_3 + k_4)$',
         transform=ax4.transAxes, fontsize=12, color=COLOR_RK4,
         bbox=dict(boxstyle='round,pad=0.5', facecolor='#21262d',
                   edgecolor=COLOR_RK4, alpha=0.9))

plt.tight_layout()
ruta_fig4 = os.path.join(output_dir, "04_rk4_pendientes.png")
fig4.savefig(ruta_fig4, dpi=200, bbox_inches='tight')
plt.show()
print(f"  [✓] Gráfica 4 guardada en: {ruta_fig4}")

# =====================================================================
# GRÁFICA 5: Convergencia — Error máximo vs tamaño de paso h
# =====================================================================
# Verificamos numéricamente que:
#   Error_Euler ~ O(h)   → pendiente 1 en log-log
#   Error_RK2   ~ O(h²)  → pendiente 2 en log-log
#   Error_RK4   ~ O(h⁴)  → pendiente 4 en log-log

pasos_h = [2.0, 1.0, 0.5, 0.25, 0.125, 0.0625]
err_max_euler = []
err_max_rk2   = []
err_max_rk4   = []

for h_test in pasos_h:
    # Solo considerar h que no violen estabilidad de Euler (kd*h < 2)
    if kd * h_test >= 2.0:
        # Euler y RK2 serían inestables, marcar como NaN
        err_max_euler.append(np.nan)
        err_max_rk2.append(np.nan)
    else:
        # Ejecutar Euler
        t_e, X_e = metodo_euler(f, 0, X0, h_test, t_final)
        exactos_e = solucion_exacta(t_e)
        errores_e = np.abs(X_e - exactos_e) / np.abs(exactos_e) * 100
        err_max_euler.append(np.max(errores_e[1:]))  # ignorar t=0

        # Ejecutar RK2
        t_r2, X_r2 = metodo_rk2_heun(f, 0, X0, h_test, t_final)
        exactos_r2 = solucion_exacta(t_r2)
        errores_r2 = np.abs(X_r2 - exactos_r2) / np.abs(exactos_r2) * 100
        err_max_rk2.append(np.max(errores_r2[1:]))

    # RK4 siempre estable para estos h
    t_r4, X_r4 = metodo_rk4(f, 0, X0, h_test, t_final)
    exactos_r4 = solucion_exacta(t_r4)
    errores_r4 = np.abs(X_r4 - exactos_r4) / np.abs(exactos_r4) * 100
    err_max_rk4.append(np.max(errores_r4[1:]))

fig5, ax5 = plt.subplots(figsize=(10, 7))

ax5.loglog(pasos_h, err_max_euler, 'v-', color=COLOR_EULER, linewidth=2,
           markersize=8, label='Euler — $O(h)$')
ax5.loglog(pasos_h, err_max_rk2, '^-', color=COLOR_RK2, linewidth=2,
           markersize=8, label='RK2 — $O(h^2)$')
ax5.loglog(pasos_h, err_max_rk4, 's-', color=COLOR_RK4, linewidth=2,
           markersize=8, label='RK4 — $O(h^4)$')

# Líneas de referencia de pendiente teórica
h_ref = np.array([0.0625, 2.0])
# O(h): pendiente 1
ax5.loglog(h_ref, 30 * h_ref**1, ':', color=COLOR_EULER, alpha=0.4, linewidth=1)
# O(h²): pendiente 2
ax5.loglog(h_ref, 15 * h_ref**2, ':', color=COLOR_RK2, alpha=0.4, linewidth=1)
# O(h⁴): pendiente 4
ax5.loglog(h_ref, 8 * h_ref**4, ':', color=COLOR_RK4, alpha=0.4, linewidth=1)

ax5.set_xlabel('Tamaño de paso h (días)', fontsize=13, fontweight='bold')
ax5.set_ylabel('Error relativo máximo (%)', fontsize=13, fontweight='bold')
ax5.set_title('Orden de Convergencia de cada Método\n'
              'Pendiente en log-log = orden del método',
              fontsize=15, fontweight='bold', pad=15)

ax5.legend(loc='upper left', fontsize=11, framealpha=0.9)
ax5.grid(True, which='both', alpha=0.3)

plt.tight_layout()
ruta_fig5 = os.path.join(output_dir, "05_convergencia.png")
fig5.savefig(ruta_fig5, dpi=200, bbox_inches='tight')
plt.show()
print(f"  [✓] Gráfica 5 guardada en: {ruta_fig5}")

# =====================================================================
# GRÁFICA 6: Datos experimentales vs modelo (discrepancia)
# =====================================================================
# Los datos reales NO siguen un decaimiento exponencial puro porque:
# - Bacterias VBNC (viable no cultivable): dejan de formar colonias
#   pero no están muertas.
# - El modelo exponencial simple no captura este comportamiento bifásico.

fig6, (ax6a, ax6b) = plt.subplots(1, 2, figsize=(14, 6))

# Panel izquierdo: escala log
ax6a.semilogy(t_fino, X_exacta_fino, color=COLOR_EXACTA, linewidth=2,
              label='Modelo: $X_0 e^{-k_d t}$')
ax6a.semilogy(t_exp, X_exp, 'D-', color=COLOR_EXP, linewidth=2,
              markersize=10, markeredgecolor='white', markeredgewidth=1.5,
              label='Datos Experimentales')
ax6a.semilogy(t_rk4, X_rk4, 's:', color=COLOR_RK4, linewidth=1,
              markersize=4, alpha=0.7, label='RK4 (h=1)')

ax6a.set_xlabel('Tiempo (días)', fontsize=12, fontweight='bold')
ax6a.set_ylabel('X (UFC/ml) — escala log', fontsize=12, fontweight='bold')
ax6a.set_title('Escala Logarítmica', fontsize=13, fontweight='bold')
ax6a.legend(fontsize=9, framealpha=0.9)
ax6a.grid(True, which='both', alpha=0.3)
ax6a.set_xlim(-0.5, 15.5)

# Panel derecho: escala lineal (muestra que el modelo subestima al inicio)
ax6b.plot(t_fino, X_exacta_fino, color=COLOR_EXACTA, linewidth=2,
          label='Modelo: $X_0 e^{-k_d t}$')
ax6b.plot(t_exp, X_exp, 'D-', color=COLOR_EXP, linewidth=2,
          markersize=10, markeredgecolor='white', markeredgewidth=1.5,
          label='Datos Experimentales')

ax6b.set_xlabel('Tiempo (días)', fontsize=12, fontweight='bold')
ax6b.set_ylabel('X (UFC/ml) — escala lineal', fontsize=12, fontweight='bold')
ax6b.set_title('Escala Lineal', fontsize=13, fontweight='bold')
ax6b.legend(fontsize=9, framealpha=0.9)
ax6b.grid(True, alpha=0.3)
ax6b.set_xlim(-0.5, 15.5)

fig6.suptitle('Discrepancia: Modelo Exponencial vs Datos Reales\n'
              '(Los datos no siguen un decaimiento exponencial puro — efecto VBNC)',
              fontsize=14, fontweight='bold', y=1.02)

plt.tight_layout()
ruta_fig6 = os.path.join(output_dir, "06_modelo_vs_datos.png")
fig6.savefig(ruta_fig6, dpi=200, bbox_inches='tight')
plt.show()
print(f"  [✓] Gráfica 6 guardada en: {ruta_fig6}")

# =====================================================================
# RESUMEN FINAL
# =====================================================================
print("\n" + "=" * 65)
print("  RESUMEN FINAL — PROYECTO DE MÉTODOS NUMÉRICOS")
print("=" * 65)
print(f"""
  EDO:    dX/dt = -kd * X(t)
  kd:     {kd:.6f} día⁻¹
  X(0):   {X0:.2e} UFC/ml
  Paso h: {h} día
  kd·h:   {kd*h:.4f}

  ┌──────────┬────────┬──────────────────────┐
  │ Método   │ Orden  │ Error máx en t=15    │
  ├──────────┼────────┼──────────────────────┤
  │ Euler    │ O(h)   │ {abs(X_euler[15] - solucion_exacta(15))/solucion_exacta(15)*100:>18.2f}% │
  │ RK2      │ O(h²)  │ {abs(X_rk2[15] - solucion_exacta(15))/solucion_exacta(15)*100:>18.2f}% │
  │ RK4      │ O(h⁴)  │ {abs(X_rk4[15] - solucion_exacta(15))/solucion_exacta(15)*100:>18.2f}% │
  └──────────┴────────┴──────────────────────┘

  Conclusión:
  RK4 (orden 4) ofrece la mejor precisión con error < 6%%.
  Euler (orden 1) acumula error de ~100%% con h=1.
  RK2 (orden 2) sobreestima la solución con error > 474%%.
  Para esta EDO con kd·h = 0.748, RK4 es el método óptimo.
""")
print("=" * 65)
print("  Gráficas guardadas en:", output_dir)
print("=" * 65)
